from flask import send_file, request, jsonify
from models.bien_model import BienModel
from models.movimiento_model import MovimientoModel
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from io import BytesIO
from io import BytesIO
from datetime import datetime
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

class ReporteController:
    @staticmethod
    def get_options():
        options = BienModel.get_report_options()
        return jsonify(options)

    @staticmethod
    def get_recent_activity():
        try:
            activities = MovimientoModel.get_recent_activity()
            
            # Format for frontend
            formatted_activities = []
            for activity in activities:
                # Date formatting
                fecha = activity['fecha']
                if fecha:
                    # If it's a date object, format it. If it's a string, leave it (or parse and reformat if needed)
                    # Assuming it might be a date object from pymysql
                    if not isinstance(fecha, str):
                        fecha = fecha.strftime('%d %b %Y')
                
                # Status color logic
                status_color = 'bg-green-100 text-green-700'
                if activity['accion'] == 'Traslado':
                    status_color = 'bg-yellow-100 text-yellow-700'
                elif activity['accion'] == 'Baja':
                    status_color = 'bg-red-100 text-red-700'
                elif activity['accion'] == 'Mantenimiento':
                    status_color = 'bg-blue-100 text-blue-700'
                
                formatted_activities.append({
                    "bien": activity['bien'],
                    "accion": activity['accion'],
                    "usuario": activity['usuario'] or 'Sin asignar',
                    "fecha": fecha,
                    "status": activity['status'],
                    "statusColor": status_color
                })
                
            return jsonify(formatted_activities)
        except Exception as e:
            print(f"Error getting recent activity: {e}")
            return jsonify([]), 500

    @staticmethod
    def get_movements_chart():
        try:
            data = MovimientoModel.get_movements_by_month()
            
            # Initialize all months with 0
            months_data = {i: 0 for i in range(1, 13)}
            
            # Fill with actual data
            for row in data:
                months_data[row['mes']] = row['total']
            
            # Labels
            labels = ['Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun', 'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']
            
            # Calculate max for height scaling
            max_value = max(months_data.values())
            if max_value == 0:
                max_value = 1 # Avoid division by zero
                
            chart_data = []
            for i in range(1, 13):
                value = months_data[i]
                # Height is percentage of max value
                height = (value / max_value) * 100
                
                chart_data.append({
                    "label": labels[i-1],
                    "value": value,
                    "height": round(height, 1)
                })
                
            return jsonify(chart_data)
        except Exception as e:
            print(f"Error getting chart data: {e}")
            return jsonify([]), 500

    @staticmethod
    def get_detalles_options():
        try:
            detalles = BienModel.get_unique_detalles()
            return jsonify(detalles)
        except Exception as e:
            print(f"Error getting detalles options: {e}")
            return jsonify([]), 500

    @staticmethod
    def generar_pdf_bienes_estado():
        try:
            data = request.json
            detalle = data.get('detalle')
            estado = data.get('estado')
            
            bienes = BienModel.get_for_pdf_report(detalle, estado)
            
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
            elements = []
            
            styles = getSampleStyleSheet()
            styles.add(ParagraphStyle(name='CenterTitle', parent=styles['Heading1'], alignment=1))
            
            # Title
            title_text = "Reporte de Bienes"
            if detalle:
                title_text += f" - {detalle}"
            if estado:
                title_text += f" ({estado})"
            
            elements.append(Paragraph(title_text, styles['CenterTitle']))
            elements.append(Spacer(1, 20))
            
            # Table Data
            table_data = [['Código', 'Descripción', 'Marca', 'Modelo', 'Estado', 'Ubicación']]
            
            for bien in bienes:
                row = [
                    bien['codigo_patrimonio'] or bien.get('codigo_interno', '-'),
                    Paragraph(bien['descripcion'], styles['Normal']), # Wrap text
                    bien['marca'],
                    bien['modelo'],
                    bien['estado'],
                    Paragraph(bien['ubicacion_actual'] or 'Sin asignar', styles['Normal'])
                ]
                table_data.append(row)
            
            if len(bienes) == 0:
                elements.append(Paragraph("No se encontraron bienes con los filtros seleccionados.", styles['Normal']))
            else:
                # Table Style
                table = Table(table_data, colWidths=[80, 150, 70, 70, 60, 100])
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('ALIGN', (1, 1), (1, -1), 'LEFT'), # Align description to left
                    ('ALIGN', (5, 1), (5, -1), 'LEFT'), # Align location to left
                ]))
                elements.append(table)
            
            # Footer info
            elements.append(Spacer(1, 20))
            elements.append(Paragraph(f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
            
            doc.build(elements)
            buffer.seek(0)
            
            return send_file(buffer, download_name="reporte_bienes.pdf", as_attachment=True, mimetype='application/pdf')
            
        except Exception as e:
            print(f"Error generating PDF: {e}")
            return jsonify({"error": str(e)}), 500

    @staticmethod
    def generar_excel_bienes_responsable():
        data = request.json
        responsable = data.get('responsable', '')
        area = data.get('area', '')
        sede = data.get('sede', '')
        
        filters = {
            'responsable': responsable,
            'area': area
        }
        
        bienes = BienModel.get_for_report(filters)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte de Bienes"
        
        # Page Setup
        # Page Setup
        ws.page_setup.paperSize = 9 # A4
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0 # Automatic height
        
        # Styles
        thin_border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))
        
        header_font = Font(bold=True, size=10)
        title_font = Font(bold=True, size=12)
        
        # Title
        ws.merge_cells('A1:H1')
        ws['A1'] = "BIENES IDENTIFICADOS COMO SOBRANTES DE INVENTARIOS DE EJERC. ANTERIORES"
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        
        # Info Rows
        ws['B3'] = "SEDE:"
        ws['C3'] = sede
        ws['B4'] = "AREA:"
        ws['C4'] = area
        ws['B5'] = "RESPONSABLE:"
        ws['C5'] = responsable
        
        for row in range(3, 6):
            ws[f'B{row}'].font = header_font
        
        # Table Headers
        headers = [
            "CODIGO PATRIMONIAL", "CODIGO INTERNO", "DETALLE DEL BIEN", 
            "CARACTERISTICAS", "UBICACIÓN", "ESTADO", "CANTIDAD", "IMPORTE"
        ]
        
        header_row = 8
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num, value=header)
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
            
        # Data
        start_row = 9
        for i, bien in enumerate(bienes):
            row_num = start_row + i
            
            # Map fields
            caracteristicas = []
            if bien['marca']: caracteristicas.append(f"MARCA: {bien['marca']}")
            if bien['modelo']: caracteristicas.append(f"MOD: {bien['modelo']}")
            if bien['dimension']: caracteristicas.append(f"DIMS: {bien['dimension']}")
            if bien['color']: caracteristicas.append(f"COLOR: {bien['color']}")
            if bien['detalle_bien']: caracteristicas.append(bien['detalle_bien'])
            
            caracteristicas_str = ", ".join(caracteristicas)
            
            estado_map = {'BUENO': 'B', 'REGULAR': 'R', 'MALO': 'M'}
            estado_short = estado_map.get(bien['estado'], bien['estado'])
            
            row_data = [
                bien['codigo_patrimonio'],
                bien['codigo_interno'],
                caracteristicas_str,
                bien['descripcion'],
                bien['ubicacion_nombre'],
                estado_short,
                1, # Cantidad
                0.00 # Importe
            ]
            
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center', wrap_text=True)
                
                if col_num == 8: # Importe
                    cell.number_format = '0.00'
        
        # Column Widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 8
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 10
        
        # Date
        months = ["ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO", "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"]
        now = datetime.now()
        date_str = f"HUÁNUCO, {now.day} DE {months[now.month-1]} DEL {now.year}"
        
        last_row = start_row + len(bienes) + 2
        ws.merge_cells(f'C{last_row}:F{last_row}')
        ws[f'C{last_row}'] = date_str
        ws[f'C{last_row}'].alignment = Alignment(horizontal='right')
        
        # Signatures
        sig_row = last_row + 4
        
        # Signature 1
        ws.merge_cells(f'B{sig_row}:C{sig_row}')
        ws[f'B{sig_row}'] = "JEFE INMEDIATO"
        ws[f'B{sig_row}'].border = Border(top=Side(style='thin'))
        ws[f'B{sig_row}'].alignment = Alignment(horizontal='center')
        
        # Signature 2
        ws.merge_cells(f'E{sig_row}:G{sig_row}')
        ws[f'E{sig_row}'] = "CONTROL PATRIMONIAL"
        ws[f'E{sig_row}'].border = Border(top=Side(style='thin'))
        ws[f'E{sig_row}'].alignment = Alignment(horizontal='center')

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(output, download_name="reporte_bienes.xlsx", as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
